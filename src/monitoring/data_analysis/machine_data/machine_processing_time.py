import glob
import os
from datetime import datetime

import numpy as np
import pandas as pd
from matplotlib import pyplot as plt
from openpyxl import load_workbook
from openpyxl.worksheet.table import TableStyleInfo, Table

from src import MACHINE_STATISTICS, MACHINE_STATISTICS_GRAPH
from src.monitoring.data_analysis.creating_machine_during_simulation_dict import CreatingMachineDuringSimulationDict


class MachineProcessingTime:
    every_machine_during_simulation_data: dict[str, list[dict]]
    machine_identification_str_list: list[str]
    machine_statistics_per_product_data: pd.DataFrame
    total_machine_statistics_data: pd.DataFrame

    def __init__(self, creating_machine_during_simulation_dict: CreatingMachineDuringSimulationDict):
        self.creating_machine_during_simulation_dict = creating_machine_during_simulation_dict

        self.every_machine_during_simulation_data = self.creating_machine_during_simulation_dict.every_machine_during_simulation_data
        self.machine_identification_str_list = self.creating_machine_during_simulation_dict.every_machine_identification_str_list

        self.machine_statistics_per_product_data = pd.DataFrame()
        self.total_machine_statistics_data = pd.DataFrame()

        self.clear_machine_statistics_files()

        self.get_production_machine_data()

        self.plot_mean_stddev_variance()
        self.save_machine_statistics_table()

    def get_production_machine_data(self):
        machine_stats_per_product_list = []
        total_machine_stats_list = []

        for identification_str in self.machine_identification_str_list:
            machine_dict_list = self.every_machine_during_simulation_data.get(identification_str, [])

            unique_product_list = [
                p for p in self.get_list_of_produced_product_group_for_machine(machine_dict_list) if p
            ]

            machine_total_processing_time_list = []
            machine_total_waiting_input_time_list = []
            machine_total_waiting_output_time_list = []

            for product in unique_product_list:
                # Analyse je Produktgruppe
                processing_times = self.get_list_processing_time_per_product(machine_dict_list, product)
                input_waiting_times = self.get_list_waiting_time_machine_input_empty(machine_dict_list, product)
                output_waiting_times = self.get_list_waiting_time_machine_output_full(machine_dict_list, product)

                machine_total_processing_time_list.extend(processing_times)
                machine_total_waiting_input_time_list.extend(input_waiting_times)
                machine_total_waiting_output_time_list.extend(output_waiting_times)

                machine_stats_per_product_list.append({
                    "machine": identification_str,
                    "product": product,
                    "number_of_produced_product": len(processing_times),

                    "mean_processing_time": self.get_mean_duration(processing_times),
                    "std_dev_processing_time": self.get_std_dev_duration(processing_times),
                    "variance_processing_time": self.get_variance_duration(processing_times),

                    "mean_waiting_input_time": self.get_mean_duration(input_waiting_times),
                    "std_dev_waiting_input_time": self.get_std_dev_duration(input_waiting_times),
                    "variance_waiting_input_time": self.get_variance_duration(input_waiting_times),

                    "mean_waiting_output_time": self.get_mean_duration(output_waiting_times),
                    "std_dev_waiting_output_time": self.get_std_dev_duration(output_waiting_times),
                    "variance_waiting_output_time": self.get_variance_duration(output_waiting_times),
                })

            total_machine_stats_list.append({
                "machine": identification_str,
                "number_of_produced_product": len(machine_total_processing_time_list),

                "mean_processing_time": self.get_mean_duration(machine_total_processing_time_list),
                "std_dev_processing_time": self.get_std_dev_duration(machine_total_processing_time_list),
                "variance_processing_time": self.get_variance_duration(machine_total_processing_time_list),

                "mean_waiting_input_time": self.get_mean_duration(machine_total_waiting_input_time_list),
                "std_dev_waiting_input_time": self.get_std_dev_duration(machine_total_waiting_input_time_list),
                "variance_waiting_input_time": self.get_variance_duration(machine_total_waiting_input_time_list),

                "mean_waiting_output_time": self.get_mean_duration(machine_total_waiting_output_time_list),
                "std_dev_waiting_output_time": self.get_std_dev_duration(machine_total_waiting_output_time_list),
                "variance_waiting_output_time": self.get_variance_duration(machine_total_waiting_output_time_list),
            })

        self.machine_statistics_per_product_data = pd.DataFrame(machine_stats_per_product_list)
        self.total_machine_statistics_data = pd.DataFrame(total_machine_stats_list)

    def get_list_of_produced_product_group_for_machine(self, machine_data_during_simulation: list[dict]) -> list[str]:
        """
        Input a list of dictionaries of one machine with every status change and timestamp. Creating a list of every
        product this machine has produced in the production
        :return list of str (product group: ex. "ProductGroup.THREE.1)"""
        unique_products = set()

        for machine_dict in machine_data_during_simulation:
            for entity in machine_dict.get("entities", []):
                if entity["entity_type"] == "Machine":
                    material = entity["entity_data"]["working_status"]["producing_production_material"]
                    if material:
                        unique_products.add(material)

        return list(unique_products)

    def get_list_processing_time_per_product(self, machine_data_during_simulation: list[dict], product_group: str) -> \
            list[int]:
        """Calculates the processing times of a specific product, without being disturbed by other
            intermediate products."""
        processing_times = []
        start_time = None

        for entry in machine_data_during_simulation:
            timestamp = entry.get("timestamp")
            for entity in entry.get("entities", []):
                status = entity.get("entity_data", {}).get("working_status", {})
                current_material = status.get("producing_production_material")
                producing_item = status.get("producing_item")
                storage_status = status.get("storage_status")

                if current_material == product_group and producing_item:
                    if start_time is None:
                        start_time = timestamp
                else:
                    if start_time is not None:
                        processing_times.append(timestamp - start_time)
                        start_time = None


        # Falls die Verarbeitung am Ende noch läuft
        if start_time is not None:
            last_timestamp = machine_data_during_simulation[-1]["timestamp"]
            processing_times.append(last_timestamp - start_time)

        return processing_times

    def get_list_waiting_time_machine_input_empty(self, machine_data_during_simulation: list[dict],
                                                  product_group: str) -> list[int]:
        """Calculates waiting times due to a lack of input material for a specific product."""
        waiting_times = []
        start_time = None

        for entry in machine_data_during_simulation:
            timestamp = entry.get("timestamp")
            for entity in entry.get("entities", []):
                status = entity.get("entity_data", {}).get("working_status", {})
                current_material = status.get("producing_production_material")
                process_status = status.get("process_status")
                storage_status = status.get("storage_status")

                if (
                        current_material == product_group and
                        process_status == "producing is paused" and
                        storage_status == "input storage is empty"
                ):
                    if start_time is None:
                        start_time = timestamp
                else:
                    if start_time is not None:
                        waiting_times.append(timestamp - start_time)
                        start_time = None

        if start_time is not None:
            last_timestamp = machine_data_during_simulation[-1]["timestamp"]
            waiting_times.append(last_timestamp - start_time)

        return waiting_times

    def get_list_waiting_time_machine_output_full(self, machine_data_during_simulation: list[dict],
                                                  product_group: str) -> list[int]:
        """Calculates waiting times due to full output storage for a specific product"""
        waiting_times = []
        start_time = None

        for entry in machine_data_during_simulation:
            timestamp = entry.get("timestamp")
            for entity in entry.get("entities", []):
                status = entity.get("entity_data", {}).get("working_status", {})
                current_material = status.get("producing_production_material")
                storage_status = status.get("storage_status")

                if (
                        current_material == product_group and
                        storage_status == "output storage is full"
                ):
                    if start_time is None:
                        start_time = timestamp
                else:
                    if start_time is not None:
                        waiting_times.append(timestamp - start_time)
                        start_time = None

        if start_time is not None:
            last_timestamp = machine_data_during_simulation[-1]["timestamp"]
            waiting_times.append(last_timestamp - start_time)

        return waiting_times

    def get_mean_duration(self, durations: list[int]) -> float:
        """Input: durations
        :return arithmetic mean of the durations (e.g., processing time, input wait time, output wait time)"""
        if not durations:
            return 0.0
        mean = sum(durations) / len(durations)
        return mean

    def get_variance_duration(self, durations: list[int]) -> float:
        """Input: durations
        :return variance of the durations (e.g., processing time, input wait time, output wait time)"""
        if not durations:
            return 0.0
        mean = sum(durations) / len(durations)
        variance = sum((x - mean) ** 2 for x in durations) / len(durations)
        return variance

    def get_std_dev_duration(self, durations: list[int]) -> float:
        """Input: durations
        :return standard deviation of the durations (e.g., processing time, input wait time, output wait time)"""
        variance = self.get_variance_duration(durations)
        std_dev = variance ** 0.5
        return std_dev

    def plot_mean_stddev_variance(self):
        # Extract machines and necessary columns from the DataFrame
        machines = self.total_machine_statistics_data["machine"]
        bar_width = 0.25
        index = np.arange(len(self.total_machine_statistics_data))

        # Create the 'Mean, Std Dev, and Variance' for Processing Time
        self._plot_time_stats(machines, index, 'processing_time', 'Bearbeitungszeit', 'Bearbeitungszeit', 0)

        # Create the 'Mean, Std Dev, and Variance' for Waiting Input Time
        self._plot_time_stats(machines, index, 'waiting_input_time', 'Wartezeit Input', 'Input', 1)

        # Create the 'Mean, Std Dev, and Variance' for Waiting Output Time
        self._plot_time_stats(machines, index, 'waiting_output_time', 'Wartezeit Output', 'Output', 2)

    def _plot_time_stats(self, machines, index, time_type, time_title, label_suffix, shift):
        """
        Create a bar chart for the given time statistics (mean, std dev, variance).
        :param machines: List of machine identifiers
        :param index: X-axis position of the bars
        :param time_type: Time type (processing_time, waiting_input_time, waiting_output_time)
        :param time_title: Title for the time stats chart (e.g., 'Processing Time')
        :param label_suffix: Label suffix to be used in the plot (e.g., 'Processing Time', 'Input', 'Output')
        :param shift: The shift for the x-axis to avoid overlapping bars
        """
        # Prepare figure for the bar chart
        plt.figure(figsize=(12, 6))

        # Adjust the index for plotting
        bar_positions = index + shift * 0.4  # Calculate the new X-axis positions for the bars

        # Plot Mean, Std Dev, and Variance for the current time type
        mean_bars = plt.bar(bar_positions, self.total_machine_statistics_data[f"mean_{time_type}"], 0.25,
                            label=f"Mittelwert {time_title}", color=(113 / 255, 28 / 255, 55 / 255))
        std_dev_bars = plt.bar(bar_positions + 0.25, self.total_machine_statistics_data[f"std_dev_{time_type}"], 0.25,
                               label=f"Standardabweichung {time_title}", color=(161 / 255, 204 / 255, 201 / 255))
        if time_type == "processing_time":
            variance_bars = plt.bar(bar_positions + 0.5, self.total_machine_statistics_data[f"variance_{time_type}"],
                                    0.25,
                                    label=f"Varianz {time_title}", color=(219 / 255, 203 / 255, 150 / 255))
            self._add_values_on_bars(variance_bars)

        # Add the value labels on top of each bar
        self._add_values_on_bars(mean_bars)
        self._add_values_on_bars(std_dev_bars)

        # Customize the plot
        plt.xlabel('Maschinen')
        plt.ylabel('Zeit (Sekunden)')
        plt.title(f'Mittelwert, Standardabweichung und Varianz von {time_title}')
        plt.xticks(bar_positions + 0.25, machines, rotation=90)  # Align X-ticks with the center of the bars
        plt.legend()
        plt.tight_layout()

        # Save the plot (this will overwrite the file if it already exists)
        file_path = os.path.join(MACHINE_STATISTICS_GRAPH, f"{time_type}_stats.png")
        plt.savefig(file_path)  # No need for overwrite=True, it will automatically overwrite if the file exists
        plt.close()

    def _add_values_on_bars(self, bars):
        """
        Add values on top of each bar in a bar chart.
        :param bars: The bars object returned by plt.bar()
        """
        for bar in bars:
            yval = bar.get_height()
            plt.text(bar.get_x() + bar.get_width() / 2, yval, round(yval, 2),
                     ha='center', va='bottom', rotation=90, fontsize=9)

    def save_machine_statistics_table(self, filename: str = "Maschinendaten.xlsx"):
        # Check if MACHINE_STATISTICS directory exists, otherwise create it
        os.makedirs(MACHINE_STATISTICS, exist_ok=True)
        filepath = os.path.join(MACHINE_STATISTICS, filename)

        # Step 1: Save both DataFrames to an Excel file with different sheets
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # Save the two DataFrames in their respective sheets
            self.machine_statistics_per_product_data.to_excel(writer, index=False,
                                                              sheet_name='Machine Stats Per Product')
            self.total_machine_statistics_data.to_excel(writer, index=False, sheet_name='Total Machine Stats')

        # Step 2: Load the Excel file using openpyxl
        workbook = load_workbook(filepath)

        # Step 3: Skip table formatting (removed _apply_table_formatting)

        # Step 4: Save the Excel file without additional formatting
        workbook.save(filepath)

    def _apply_table_formatting(self, worksheet):
        # Generiere einen eindeutigen Tabellennamen, z. B. mit einem Zeitstempel
        unique_table_name = "MachineStatsTable_" + datetime.now().strftime("%Y%m%d%H%M%S")

        # Erstelle die Tabelle mit dem einzigartigen Namen
        max_row = worksheet.max_row
        max_col = worksheet.max_column
        last_col_letter = worksheet.cell(row=1, column=max_col).column_letter
        table_range = f"A1:{last_col_letter}{max_row}"

        # Erstelle die Tabelle
        table = Table(displayName=unique_table_name, ref=table_range)

        # Formatieren der Tabelle
        style = TableStyleInfo(
            name="TableStyleMedium9",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False
        )
        table.tableStyleInfo = style
        worksheet.add_table(table)

    def clear_machine_statistics_files(self):
        # Define the paths to the directories
        machine_statistics_dir = MACHINE_STATISTICS
        machine_statistics_graph_dir = MACHINE_STATISTICS_GRAPH

        # Delete all .xlsx files in MACHINE_STATISTICS directory
        xlsx_files = glob.glob(os.path.join(machine_statistics_dir, "*.xlsx"))
        for file in xlsx_files:
            try:
                os.remove(file)
                print(f"Deleted {file}")
            except Exception as e:
                print(f"Error deleting {file}: {e}")

        # Delete all .png files in MACHINE_STATISTICS_GRAPH directory
        png_files = glob.glob(os.path.join(machine_statistics_graph_dir, "*.png"))
        for file in png_files:
            try:
                os.remove(file)
                print(f"Deleted {file}")
            except Exception as e:
                print(f"Error deleting {file}: {e}")
